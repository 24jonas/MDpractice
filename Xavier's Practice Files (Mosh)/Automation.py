import openpyxl as xl
from pathlib import Path
from openpyxl.chart import BarChart, Reference

path1 = Path("CHEM-325 R4 W3 RF_Curve_TIPA.xlsx")
#print(path1.exists())

wb = xl.load_workbook("CHEM-325 R4 W3 RF_Curve_TIPA.xlsx")
sheet = wb["RF_Curve_TIPA"]
cell = sheet.cell(1,1)      # .cell(row,column)
# Alternatively cell = sheet['a1']
#print(cell.value)       # This somehow worked.

for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        cell = sheet.cell(i,j)
        print(cell.value)       # Again, this somehow worked.

for i in range(1,3):
    for j in range(1, sheet.max_column + 1):
        cell = sheet.cell(i,j)
        print(cell.value * 0.9)     # Worked

for i in range(1,3):
    cell = sheet.cell(i,1)
    x = cell.value * 0.1
    y = sheet.cell(i, 2)
    y.value = x             # This also worked.

# wb.save("2nd CHEM-325 R4 W3 RF_Curve_TIPA.xlsx")

values = Reference(sheet, min_row = 1, max_row = 2, min_col = 1, max_col = 1)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, "q4")        # Worked

# wb.save("3rd CHEM-325 R4 W3 RF_Curve_TIPA.xlsx")
